# Load template dataframe (or just receive it in constructor?)
# Find header row? Or just append rows to document and assume?
# Append each sample, setting each sample value in the appropriate column
# Write out the template to disk?
from openpyxl import load_workbook, Workbook
from .freezeman_config import HEADERS


class FHSSampleSubmissionTemplate:
    def __init__(self, template_path: str):
        try:
            self.workbook: Workbook = load_workbook(filename=template_path)
            self.worksheet = self.workbook["SampleSubmission"]
            if self.worksheet is None:
                raise RuntimeError(
                    "'SampleSubmission' worksheet not found in fms template"
                )
        except BaseException as err:
            print("Failed to load fms sample submission template")
            print(err)
            print("\n")
            raise err

        self.header_row_index, self.column_index_map = self._locate_headers()
        if self.column_index_map is None:
            raise RuntimeError(
                "Could not locate headers in fms sample submission template"
            )

    def _locate_headers(self):
        """Locate the row containing the column headers and build a map, mapping the
        column header key to the index of the column in the sheet."""

        row_counter = 0
        column_index_map = None
        for row in self.worksheet.iter_rows(max_row=20, max_col=50, values_only=True):
            row_counter += 1
            column_index_map = {}

            # Strip any leading or trailing spaces in header names.
            trimmed_row = list(map(lambda e: e.strip() if e is not None else e, row))

            for (header_key, header_value) in HEADERS.items():
                if header_value in trimmed_row:
                    column_index = (
                        trimmed_row.index(header_value) + 1
                    )  # add 1 for one-based indexing
                    column_index_map[header_key] = column_index
                else:
                    column_index_map = None
                    break

            # If all of the columns were found, return the header row index
            # and the mapping of header keys to column indices
            if column_index_map is not None:
                return (row_counter, column_index_map)

        return (None, None)

    def append_samples(self, samples):
        if self.header_row_index is None:
            raise RuntimeError('Cannot output samples - header_row_index is not defined.')

        row_index = self.header_row_index + 1
        for sample in samples:
            self._write_sample_row(row_index, sample)
            row_index += 1

    def _write_sample_row(self, row_index, sample):
        if self.column_index_map is None:
            raise RuntimeError('Cannot output samples - column_index_map is not defined.')

        # Convert sample SimpleNamespace to a regular dictionary
        sample_dictionary = vars(sample)

        # For each header key, write the corresponding value from the sample,
        # if the value is not None.
        for header_key in HEADERS:
            value = sample_dictionary[header_key]
            if value is not None:
                column_index = self.column_index_map[header_key]
                if column_index is not None:
                    self.worksheet.cell(
                        row=row_index, column=column_index
                    ).value = value

    def write_to_file(self, file_path):
        self.workbook.save(filename=file_path)
