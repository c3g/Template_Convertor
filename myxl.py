from openpyxl import load_workbook


wb_path = "data/Sample_submission_v3_8_0.xlsx"

wb = load_workbook(filename=wb_path)

ws = wb["SampleSubmission"]

ws.cell(row=8, column=1, value="TEST")

# ws1 = wb.active
# ws1.title = "test openpyxl"

# for row in range(1, 10):
#     for col in range(1, 5):
#         ws1.cell(row=row, column=col, value=f"({row, col})")

dest_filename = "data/openpyxl.xlsx"
wb.save(filename=dest_filename)
